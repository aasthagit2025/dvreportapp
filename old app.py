import streamlit as st
import pandas as pd
from io import BytesIO
import numpy as np
import pyreadstat  # Required for SPSS
import re
from openpyxl.styles import PatternFill

# --------------------------------------------------
# Page Config
# --------------------------------------------------
st.set_page_config(page_title="Survey DV Engine", layout="wide")
st.title("🛡️ Survey Data Validation Engine")
# --------------------------------------------------
# 1. Validation Resources (Aligned Layout)
# --------------------------------------------------
col_left, col_right = st.columns(2)

with col_left:
    st.markdown("### 📋 Rules Template")
    st.write("Download the Excel template to define your validation logic.")
    
    def generate_template():
        return pd.DataFrame({
            "Question": ["RQ5", "RQ7", "RQ7", "RQ10_2", "RQ11"],
            "Check_Type": ["Range;Missing", "Range;Missing", "Attribute-Skip", "Range;Missing", "Multi-Select"],
            "Condition": [
                "1-2;Not Null", 
                "1-14;Not Null", 
                "IF RQ5 IN (2) THEN 11 ELSE BLANK", 
                "1-1;Not Null",
                "Min=1"
            ],
            "Severity": ["Critical", "Critical", "Critical", "Critical", "Critical"]
        })
    
    st.download_button(
        label="📥 Download Rules Template",
        data=generate_template().to_csv(index=False).encode('utf-8'),
        file_name='DV_Rules_Template.csv',
        mime='text/csv',
        use_container_width=True
    )

with col_right:
    st.markdown("### ⚙️ DV Macro Tool")
    st.write("Download the team's VBA Macro tool for advanced processing.")
    
    # Replace 'DV_Macro.xlsm' with your actual team macro filename
    try:
        with open("DV_Syntax_Macro.xlsm", "rb") as f:
            macro_data = f.read()
            
        st.download_button(
            label="📥 Download DV Syntax Macro (.xlsm)",
            data=macro_data,
            file_name="DV_Syntax_Macro.xlsm",
            mime='application/vnd.ms-excel.sheet.macroEnabled.12',
            use_container_width=True
        )
    except FileNotFoundError:
        st.warning("⚠️ Macro file 'DV_Syntax_Macro.xlsm' not found in app folder.")
        # Placeholder button to maintain alignment even if file is missing
        st.button("📥 Download DV Macro (File Missing)", disabled=True, use_container_width=True)

# --------------------------------------------------
# 2. File Uploads (Enhanced for SPSS)
# --------------------------------------------------
st.divider()

# Added format selection to handle SPSS logic separately
import_format = st.radio("Select Data Format:", ["Excel/CSV", "SPSS (.sav)"], horizontal=True)

col1, col2 = st.columns(2)

with col1:
    if import_format == "SPSS (.sav)":
        raw_file = st.file_uploader("Upload Raw Data (SAV)", type=["sav"])
    else:
        raw_file = st.file_uploader("Upload Raw Data (CSV/XLSX)", type=["csv", "xlsx"])

with col2:
    rules_file = st.file_uploader("Upload Validation Rules (CSV/XLSX)", type=["csv", "xlsx"])

if raw_file and rules_file:
    # --- DATA IMPORT LOGIC ---
    if import_format == "SPSS (.sav)":
        import pyreadstat
        # meta contains the Variable Names and Type/Labels for your Macro
        df, meta = pyreadstat.read_sav(raw_file)

        # Simple translator: If it starts with 'A' it's a String, otherwise it's Numeric
        def translate_type(var_name):
            spss_code = meta.original_variable_types.get(var_name, "F")
            return "String" if spss_code.startswith("A") else "Numeric"
        


        sync_df = pd.DataFrame({
        "Var Name": meta.column_names,
        "Type": [translate_type(n) for n in meta.column_names]
    })
        csv_data = sync_df.to_csv(index=False)
        
        # Display metadata for macro reference
        st.success("✅ SPSS Variable View extracted!")
        st.download_button(
        label="📥 Download Sync File for Macro",
        data=sync_df.to_csv(index=False).encode('utf-8'),
        file_name="macro_sync.csv",
        mime="text/csv",
        use_container_width=True
    )
    else:
        df = pd.read_csv(raw_file) if raw_file.name.endswith('.csv') else pd.read_excel(raw_file)
    
    rules_df = pd.read_csv(rules_file) if rules_file.name.endswith('.csv') else pd.read_excel(rules_file)
    
    # --- PRE-PROCESSING ---
    df.columns = df.columns.str.strip()
    resp_id_col = df.columns[0]
    df_numeric = df.apply(pd.to_numeric, errors='coerce')
    
    failed_rows = []
    error_locations = [] 
    rows_with_errors = set()


# --- 3. Validation Core Engine ---
    for _, rule in rules_df.iterrows():
        q_name = str(rule["Question"]).strip()
        checks = str(rule["Check_Type"])
        conds = str(rule["Condition"])
        severity = rule.get("Severity", "Critical")

# --- STEP 1: SMART COLUMN SELECTION (Attribute Aware) ---
        # 1. Try Exact Match First (This handles specific attributes like RQ9_1)
        target_cols = [c for c in df.columns if c.lower() == q_name.lower()]
        
        # 2. If no exact match, search for Grid children (handles whole questions like RQ9)
        if not target_cols:
            if q_name[-1].isdigit():
                pattern = re.compile(rf"^{re.escape(q_name)}(_|[a-zA-Z]).*$", re.IGNORECASE)
            else:
                pattern = re.compile(rf"^{re.escape(q_name)}(_|[a-zA-Z]|\d)+$", re.IGNORECASE)
            target_cols = [c for c in df.columns if pattern.match(c)]
        
        if not target_cols: continue

        # --- STEP 2: SKIP PARSER (Fixed for Attributes) ---
        # Default to "Required" unless a Skip condition says otherwise
        is_required = pd.Series(True, index=df.index)
        
        if "Skip" in checks:
            try:
                # Clean the condition string
                cond_upper = conds.upper()
                if "IF " in cond_upper and " THEN " in cond_upper:
                    trigger_part = cond_upper.split("IF ")[1].split(" THEN")[0]
                    
                    # Logic: IF [BASE_Q] IN (VALUES)
                    if " IN " in trigger_part:
                        base_q, val_part = trigger_part.split(" IN ")
                        base_q = base_q.strip()
                        
                        # Convert (1,2,3) or (1-5) to a Python list
                        val_str = val_part.strip().replace('(', '').replace(')', '')
                        if "-" in val_str:
                            start, end = map(int, val_str.split("-"))
                            valid_vals = list(range(start, end + 1))
                        else:
                            valid_vals = [int(x.strip()) for x in val_str.split(",")]
                        
                        actual_base = next((c for c in df.columns if c.upper() == base_q), None)
                        if actual_base:
                            # Update requirement: Must be answered ONLY IF base question is in valid values
                            is_required = df_numeric[actual_base].isin(valid_vals)
            except Exception as e:
                st.warning(f"Skip logic error in {q_name}: {e}")

        # --- STEP 3: ROW VALIDATION ---
        for idx in df.index:
            row_raw = df.loc[idx, target_cols]
            row_num = df_numeric.loc[idx, target_cols]

            # --- ADD THESE TWO LINES HERE ---
            any_ans = row_raw.notna().any() and not (row_raw.astype(str).str.strip() == "").all()
            all_ans = row_raw.notna().all() and not (row_raw.astype(str).str.strip() == "").any()
            
            # Check if user provided any answer
            any_ans = row_raw.notna().any() and not (row_raw.astype(str).str.strip() == "").all()
            
            # ATTRIBUTE SKIP CHECK: If not required but has data -> Error
            if "Skip" in checks and not is_required[idx] and any_ans:
                failed_rows.append({
                    "RespID": df.loc[idx, resp_id_col], 
                    "Question": q_name, 
                    "Issue": "Skip Violation: Should be Blank (Attribute)", 
                    "Severity": severity
                })
                rows_with_errors.add(idx)
                for col in target_cols: error_locations.append((idx, col))
                continue # Stop further checks for this row if skip is violated

            # Skip Violation
            if "Skip" in checks and not is_required[idx] and any_ans:
                failed_rows.append({"RespID": df.loc[idx, resp_id_col], "Question": q_name, "Issue": "Skip Violation", "Severity": severity})
                rows_with_errors.add(idx)
                for col in target_cols: error_locations.append((idx, col))
                continue 

            # Missing/Grid Check
            if is_required[idx]:
                if not any_ans:
                    failed_rows.append({"RespID": df.loc[idx, resp_id_col], "Question": q_name, "Issue": "Missing response", "Severity": severity})
                    rows_with_errors.add(idx)
                    for col in target_cols: error_locations.append((idx, col))
                elif not all_ans and len(target_cols) > 1:
                    failed_rows.append({"RespID": df.loc[idx, resp_id_col], "Question": q_name, "Issue": "Incomplete Grid", "Severity": severity})
                    rows_with_errors.add(idx)
                    for col in target_cols:
                        if pd.isna(df.loc[idx, col]): error_locations.append((idx, col))


            # 3. RANGE / SINGLE SELECT
            if "Range" in checks and is_required[idx] and any_ans:
                rng = re.search(r"(\d+)-(\d+)", conds)
                if rng:
                    low, high = map(int, rng.groups())
                    for col in target_cols:
                        val = row_num[col]
                        if pd.notna(val) and not (low <= val <= high):
                            failed_rows.append({"RespID": df.loc[idx, resp_id_col], "Question": col, "Issue": f"Out of Range ({low}-{high})", "Severity": severity})
                            rows_with_errors.add(idx)
                            error_locations.append((idx, col))

            # 4. MULTI-SELECT (Restored logic)
            if "Multi-Select" in checks and is_required[idx]:
                select_count = (row_num > 0).sum()
                min_v = 1
                if "Min=" in conds:
                    try: min_v = int(re.search(r"Min=(\d+)", conds).group(1))
                    except: pass
                if select_count < min_v:
                    failed_rows.append({"RespID": df.loc[idx, resp_id_col], "Question": q_name, "Issue": f"Multi-Select: {select_count} selected (Min {min_v})", "Severity": severity})
                    rows_with_errors.add(idx)
                    for col in target_cols: error_locations.append((idx, col))

            # 5. STRAIGHTLINER (Robust)
            if "Straightliner" in checks and len(target_cols) > 1 and all_ans:
                if row_data.nunique() == 1:
                    failed_rows.append({"RespID": df.loc[idx, resp_id_col], "Question": q_name, "Issue": "Straightliner detected", "Severity": severity})
                    rows_with_errors.add(idx)
                    for col in target_cols: error_locations.append((idx, col))

            # 6. RANKING (Robust Uniqueness)
            if "Ranking" in checks and is_required[idx] and any_ans:
                clean_ranks = row_num.dropna()
                if len(clean_ranks) != len(clean_ranks.unique()):
                    failed_rows.append({"RespID": df.loc[idx, resp_id_col], "Question": q_name, "Issue": "Ranking Error: Duplicates", "Severity": severity})
                    rows_with_errors.add(idx)
                    for col in target_cols: error_locations.append((idx, col))

            # 7. OPEN END JUNK (Robust)
            if "OpenEnd_Junk" in checks and is_required[idx] and any_ans:
                text_val = str(row_data.iloc[0]).lower().strip()
                min_l = 5
                if "MinLen=" in conds:
                    try: min_l = int(re.search(r"MinLen=(\d+)", conds).group(1))
                    except: pass
                junk = ["asdf", "test", "none", "na", "abc", "n/a", "nothing", "good"]
                if len(text_val) < min_l or text_val in junk or len(set(text_val)) < 3:
                    failed_rows.append({"RespID": df.loc[idx, resp_id_col], "Question": q_name, "Issue": "OE Junk or Too Short", "Severity": severity})
                    rows_with_errors.add(idx)
                    for col in target_cols: error_locations.append((idx, col))

                    #8. Constant Sum
            if "ConstantSum" in checks and is_required[idx] and any_ans:
                target_sum = 100
                if "Total=" in conds:
                    try: target_sum = float(re.search(r"Total=(\d+)", conds).group(1))
                    except: pass
                if row_num.sum() != target_sum:
                    failed_rows.append({"RespID": df.loc[idx, resp_id_col], "Question": q_name, "Issue": f"Sum {row_num.sum()} != {target_sum}", "Severity": severity})
                    rows_with_errors.add(idx)
                    for col in target_cols: error_locations.append((idx, col))

    # --------------------------------------------------
    # 4. Report Generation & Summary
    # --------------------------------------------------
    report_df = pd.DataFrame(failed_rows)
    if not report_df.empty:
        st.write(f"### 🚩 Found {len(report_df)} Validation Issues")
        st.dataframe(report_df, use_container_width=True)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            report_df.to_excel(writer, index=False, sheet_name='Error_Log')
            
            # Create Summary Tab
            summary = report_df.groupby(["Question", "Issue"]).size().reset_index(name="Error_Count")
            summary.to_excel(writer, index=False, sheet_name='Summary')
            
            # Highlighted Raw Data
            df.to_excel(writer, index=False, sheet_name='Highlighted_Data')
            ws = writer.sheets['Highlighted_Data']
            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            
            # Highlight Cells
            for r_idx, col_name in error_locations:
                c_idx = df.columns.get_loc(col_name) + 1
                ws.cell(row=r_idx + 2, column=c_idx).fill = red_fill
            
            # Highlight RespIDs
            rid_idx = df.columns.get_loc(resp_id_col) + 1
            for r_idx in rows_with_errors:
                ws.cell(row=r_idx + 2, column=rid_idx).fill = red_fill

        st.download_button("Download Full Report & Highlighting", output.getvalue(), "Final_Validation_Report.xlsx")
    else:
        st.success("✅ Your data is clean!")